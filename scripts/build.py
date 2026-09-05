#!/usr/bin/env python3
"""Build the public/ site from data/junior.json.

Outputs:
  public/index.html          — главная: все категории
  public/<category>/index.html — 5 страниц категорий
  public/assets/*            — style.css, app.js, шрифты, vendor
  public/records.{json,csv,xlsx,md,txt} — выгрузки
  public/og-image.png, sitemap.xml, robots.txt, 404.html
  public/offline.html, sw.js  — офлайн-страница и service worker (версия = хеш содержимого)
  public/splash/*.png, screenshot-*.png — стартовые экраны iOS и скриншоты манифеста
  README.md                  — описание репозитория с актуальными числами

Ключевое: данные живут в DOM. Строка таблицы несёт data-атрибуты, по которым
app.js фильтрует и сортирует, не пересобирая разметку. Клиентского рендера
таблицы нет — единственный источник правды здесь.
"""
from __future__ import annotations

import csv
import hashlib
import html
import json
import os
import re
import shutil
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import gen_icons
import gen_og
import gen_readme
import gen_screens

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data" / "junior.json"
STATIC = ROOT / "static"
PUBLIC = ROOT / "public"
ASSETS = PUBLIC / "assets"

SITE_TITLE = "Юношеские рекорды России по плаванию"
SITE_TAGLINE = (
    "Действующие юношеские рекорды России по плаванию: личные и эстафетные, "
    "бассейны 50 и 25 м. Таблица обновляется ежедневно. "
    "Скачать бесплатно: JSON, CSV, XLSX, PDF."
)
SITE_KEYWORDS = (
    "юношеские рекорды России по плаванию, юниорские рекорды плавание, юниоры, "
    "рекорды России по плаванию, плавание рекорды, вольный стиль, брасс, баттерфляй, "
    "на спине, комплексное плавание, бассейн 50м, бассейн 25м, russwimming"
)
SITE_DOMAIN = "russwimming-records-junior.borozdov.ru"

# QR на картинках ведёт не на сайт напрямую, а на переходник /go/.
# Чтобы перенаправить уже напечатанные плакаты, достаточно поменять эту строку.
QR_TARGET = "https://russwimming-records-junior.borozdov.ru/"
# Подпись автора — в полосе под шапкой и в подвале; одна константа, чтобы не разъехались
SIGNATURE = "СДЕЛАЛ <b>NIKITA BOROZDOV</b><br>МАСТЕР СПОРТА"
# REPO_URL из окружения переопределяет адрес репозитория (форки). По умолчанию —
# не пусто, а адрес из gen_readme: HTML входит в хеш sw.js, и локальная сборка
# без переменной давала бы другие страницы, чем CI, а с ними — новый app-<хеш>
# и холостой коммит «data: update records».
REPO_URL_ENV = "REPO_URL"


def repo_url() -> str:
    return (os.environ.get(REPO_URL_ENV, "").strip() or gen_readme.DEFAULT_REPO_URL).rstrip("/")


# UTM на исходящих ссылках бренда: по ним на borozdov.ru видно, сколько людей
# пришло отсюда и из какой точки страницы. Параметра три и они фиксированные —
# urlencode ради них не нужен. `&` отдаём как `&amp;`: голый амперсанд в href —
# невалидный HTML.
def brand_link(url: str, campaign: str) -> str:
    return (f"{url}?utm_source={SITE_DOMAIN}&amp;utm_medium=referral"
            f"&amp;utm_campaign={campaign}")


# Видимый ответ FAQ и FAQPage — один текст: в разметку он идёт как HTML (внутри
# живут ссылки), в JSON-LD — со снятыми тегами. Второго «чистого» поля нет
# намеренно: разъезжаться нечему, а Google требует совпадения. unescape
# обязателен — если в ответе появится &laquo;, в JSON-LD должен уехать символ,
# а не entity.
def faq_plain(answer: str) -> str:
    return html.unescape(re.sub(r"<[^>]+>", "", answer))


FRESH_DAYS = 365  # «Новое» — последние 12 месяцев от даты сборки

STROKE_LABELS = {
    "freestyle": "Вольный стиль",
    "backstroke": "На спине",
    "breaststroke": "Брасс",
    "butterfly": "Баттерфляй",
    "im": "Комплексное плавание",
    "medley_relay": "Комбинированная эстафета",
    "unknown": "Прочее",
}

# Видимый FAQ + FAQPage в JSON-LD собираются из одного списка, чтобы разметка
# всегда совпадала с контентом (требование Google). Ответы — HTML, а не текст:
# внутри живут ссылки. В JSON-LD тот же ответ уезжает через faq_plain().
FAQ_ITEMS = [
    (
        "Что такое юношеский рекорд России по плаванию?",
        "Лучший официально ратифицированный Всероссийской федерацией плавания "
        "результат российского спортсмена юниорского возраста или юниорской "
        "сборной, показанный в бассейне 50 или 25 метров.",
    ),
    (
        "Как часто обновляется таблица рекордов?",
        "Автоматически раз в сутки: данные синхронизируются с официальной "
        "таблицей юношеских рекордов на сайте russwimming.ru. Дата последнего "
        "обновления показана над таблицей.",
    ),
    (
        "Где посмотреть взрослые рекорды России?",
        "Взрослые рекорды России по плаванию ведутся отдельной таблицей — "
        'на сайте <a href="https://russwimming-records.borozdov.ru/" '
        'rel="noopener" target="_blank">russwimming-records.borozdov.ru</a>. '
        "Оба зеркала обновляются из одного источника и устроены одинаково.",
    ),
    (
        "Можно ли скачать таблицу рекордов?",
        "Да, бесплатно и без регистрации: JSON, CSV, XLSX, Markdown, TXT, "
        "а также PNG-картинка и PDF-документ таблицы.",
    ),
    (
        "Кто ведёт этот сайт?",
        "Nikita Borozdov — мастер спорта России по плаванию, стайер на "
        "дистанциях 400, 800 и 1500 метров. "
        "Сайт делается и поддерживается бесплатно: данные открыты, без рекламы "
        "и регистрации. Другие проекты и контакты — "
        f'<a href="{brand_link("https://borozdov.ru/", "faq_author")}" '
        'rel="noopener" target="_blank">borozdov.ru</a>.',
    ),
]

DOWNLOAD_FORMATS = [
    ("records.json", "JSON", "структурированные данные"),
    ("records.csv", "CSV", "Excel / Numbers"),
    ("records.xlsx", "XLSX", "книга по листам"),
    ("records.md", "MD", "таблицы Markdown"),
    ("records.txt", "TXT", "фиксированная ширина"),
]

CSV_HEADERS = [
    "Категория", "Дисциплина", "Тип", "Спортсмен / Команда", "Состав эстафеты",
    "Результат", "Результат (сек)", "Место", "Дата", "Дата (ISO)",
]


# ---------------------------------------------------------------- данные

def load_data() -> dict:
    return json.loads(DATA.read_text(encoding="utf-8"))


def display_discipline(discipline: str) -> str:
    """Убираем дублирующую пометку про бассейн — она уже есть в названии категории."""
    t = re.sub(r"\s*\(бассейн 25 м\)\s*", "", discipline)
    t = re.sub(r"\s{2,}", " ", t).strip()
    return t[:1].upper() + t[1:] if t else t


def enrich(data: dict, today: date) -> dict:
    """Дописывает к записям производные поля (с подчёркиванием) прямо в data.

    `_fresh` считаем здесь, а не в браузере: признак свежести не должен
    зависеть от часового пояса и календаря клиента.
    """
    for cat in data["categories"]:
        for r in cat["records"]:
            # бассейн берём из записи, а не из категории:
            # в смешанных эстафетах соседствуют обе воды
            r["_pool"] = "scm" if r["is_25m_pool"] else "lcm"
            r["_sex"] = cat["sex"]
            r["_cat_id"] = cat["id"]
            r["_cat_title"] = cat["title"]
            r["_title"] = display_discipline(r["discipline"])
            iso = r["date"]
            days = (today - date.fromisoformat(iso)).days if iso else None
            r["_fresh"] = days is not None and days <= FRESH_DAYS
    return data


def public_json(data: dict) -> dict:
    """Публичная выгрузка без служебных полей — records.json остаётся чистым API."""
    return {
        **{k: v for k, v in data.items() if k != "categories"},
        "categories": [
            {
                **{k: v for k, v in cat.items() if k != "records"},
                "records": [
                    {k: v for k, v in r.items() if not k.startswith("_")}
                    for r in cat["records"]
                ],
            }
            for cat in data["categories"]
        ],
    }


def all_records(data: dict) -> list[dict]:
    return [r for cat in data["categories"] for r in cat["records"]]


# ---------------------------------------------------------------- выгрузки

def record_rows_for_csv(data: dict):
    for cat in data["categories"]:
        for r in cat["records"]:
            yield [
                cat["title"],
                r["discipline"],
                "эстафета" if r["relay"] else "личное",
                r["athlete"],
                " · ".join(r["roster"]) if r["roster"] else "",
                r["result"],
                f"{r['result_seconds']:.2f}" if r["result_seconds"] is not None else "",
                r["location"],
                r["date_original"],
                r["date"] or "",
            ]


def write_csv(data: dict, out: Path) -> None:
    with out.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, dialect="excel")
        w.writerow(CSV_HEADERS)
        for row in record_rows_for_csv(data):
            w.writerow(row)


def write_xlsx(data: dict, out: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FAFAFA")
    header_fill = PatternFill("solid", fgColor="0D0D0D")
    mono_font = Font(name="Consolas")  # есть в Office на macOS и Windows
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right")

    for cat in data["categories"]:
        ws = wb.create_sheet(title=cat["title"][:31])  # sheet name limit
        ws.append(["Дисциплина", "Спортсмен / Команда", "Состав эстафеты",
                   "Результат", "Место", "Дата"])
        for c in ws[1]:
            c.font, c.fill, c.alignment = header_font, header_fill, center
        for r in cat["records"]:
            ws.append([
                r["discipline"], r["athlete"],
                " · ".join(r["roster"]) if r["roster"] else "",
                r["result"], r["location"], r["date_original"],
            ])
        for i, w in enumerate([34, 26, 42, 12, 22, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row_idx in range(2, ws.max_row + 1):  # результат — моно и вправо, как на сайте
            cell = ws.cell(row=row_idx, column=4)
            cell.font, cell.alignment = mono_font, right
        ws.freeze_panes = "A2"

    ws = wb.create_sheet(title="Все", index=0)
    ws.append(CSV_HEADERS)
    for c in ws[1]:
        c.font, c.fill, c.alignment = header_font, header_fill, center
    for row in record_rows_for_csv(data):
        ws.append(row)
    for i, w in enumerate([28, 34, 10, 24, 42, 12, 14, 22, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in (6, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font, cell.alignment = mono_font, right
    ws.freeze_panes = "A2"

    # Без этого openpyxl пишет в docProps/core.xml текущее время,
    # и файл меняет байты на каждой сборке даже при неизменных данных
    wb.properties.creator = SITE_TITLE
    wb.properties.lastModifiedBy = SITE_TITLE
    wb.properties.created = datetime(2020, 1, 1)
    wb.properties.modified = datetime(2020, 1, 1)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def write_markdown(data: dict, out: Path) -> None:
    lines = [
        f"# {SITE_TITLE}",
        "",
        f"Источник: <{data['source_url']}>  ",
        f"Обновлено: {data['fetched_at']}  ",
        f"Всего рекордов: {data['total_records']}",
        "",
    ]
    for cat in data["categories"]:
        lines += [f"## {cat['title']}", "",
                  "| Дисциплина | Спортсмен | Результат | Место | Дата |", "|---|---|---|---|---|"]
        for r in cat["records"]:
            athlete = r["athlete"]
            if r["roster"]:
                athlete += " (" + ", ".join(r["roster"]) + ")"
            lines.append(
                f"| {r['discipline']} | {athlete} | `{r['result']}` | "
                f"{r['location']} | {r['date_original']} |"
            )
        lines.append("")
    out.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_txt(data: dict, out: Path) -> None:
    """Fixed-width plain-text dump — handy for grep / terminal users."""
    entries = []
    for cat in data["categories"]:
        for r in cat["records"]:
            athlete = r["athlete"]
            if r["roster"]:
                athlete += " (" + ", ".join(r["roster"]) + ")"
            entries.append((cat["title"], r["discipline"], r["result"], athlete,
                            r["location"], r["date_original"]))

    # ширины колонок — по фактическим данным, чтобы эстафеты не ломали сетку
    disc_w = max(len(e[1]) for e in entries)
    res_w = max(len(e[2]) for e in entries)
    ath_w = max(len(e[3]) for e in entries)
    loc_w = max(len(e[4]) for e in entries)

    rows = [
        SITE_TITLE.upper(),
        f"Источник: {data['source_url']}",
        f"Обновлено: {data['fetched_at']} · всего рекордов: {data['total_records']}",
        "",
    ]
    for cat in data["categories"]:
        rows.append(f"=== {cat['title']} ===")
        for title, disc, res, athlete, loc, date_s in entries:
            if title != cat["title"]:
                continue
            rows.append(f"  {disc:<{disc_w}}  {res:>{res_w}}  {athlete:<{ath_w}}  {loc:<{loc_w}}  {date_s}")
        rows.append("")
    out.write_text("\n".join(rows) + "\n", encoding="utf-8")


def copy_static() -> None:
    ASSETS.mkdir(parents=True, exist_ok=True)
    # fonts.css в assets не копируем: @font-face инлайнится в <head> (см. fonts_css()),
    # чтобы браузер узнавал о шрифтах из первых байт HTML, а не после второго запроса
    for name in ("style.css", "app.js"):
        shutil.copy2(STATIC / name, ASSETS / name)
    stale = ASSETS / "fonts.css"
    if stale.exists():
        stale.unlink()
    for sub in ("fonts", "vendor"):
        src = STATIC / sub
        if src.exists():
            shutil.copytree(src, ASSETS / sub, dirs_exist_ok=True)


# ---------------------------------------------------------------- версии ассетов

def file_hash(path: Path, length: int = 8) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()[:length]


def asset_url(rel: str) -> str:
    """`/assets/<rel>?v=<хеш содержимого>`.

    Имена файлов не меняются, поэтому версия живёт в query. Она нужна не HTTP-кэшу
    (хостинг Cache-Control не шлёт, браузер кэширует эвристически), а service worker'у: HTML новой сборки
    ссылается на новые URL, и старый кэш их не найдёт — страница и её скрипт всегда
    из одной сборки. Тот же URL идёт в прекэш sw.js, поэтому совпадение точное.
    """
    return f"/assets/{rel}?v={file_hash(STATIC / rel)}"


def fonts_css() -> str:
    """@font-face из static/fonts.css с абсолютными версионированными путями."""
    css = (STATIC / "fonts.css").read_text(encoding="utf-8")
    return re.sub(r"url\('\./fonts/([^']+)'\)",
                  lambda m: f"url('{asset_url('fonts/' + m.group(1))}')", css)


def font_urls() -> list[str]:
    return [asset_url(f"fonts/{p.name}") for p in sorted((STATIC / "fonts").glob("*.woff2"))]


def vendor_urls() -> list[str]:
    return [f"/assets/vendor/{p.name}" for p in sorted((STATIC / "vendor").glob("*.js"))]


# ---------------------------------------------------------------- рендер таблицы

def fmt_date_ru(iso: str, orig: str) -> str:
    return ".".join(reversed(iso.split("-"))) if iso else (orig or "")


# Сортируются все пять: заголовок = кнопка с тактами ↑ → ↓ → выкл
COLUMNS = [
    ("disc", "Дисциплина"),
    ("athlete", "Спортсмен"),
    ("result", "Результат"),
    ("location", "Место"),
    ("date", "Дата"),
]


def render_table(categories: list[dict], caption: str) -> str:
    """Серверный рендер — единственный. app.js только прячет и переставляет строки.

    Явные role=* обязательны: ниже 720px CSS превращает строки в grid,
    а смена display стирает табличную семантику.
    """
    e = html.escape

    head = "".join(
        f'<th scope="col" role="columnheader" class="col-{key} th-sortable" '
        f'data-key="{key}" aria-sort="none">'
        f'<button type="button" class="th-btn">{label}'
        f'<span class="sort-ind" aria-hidden="true"></span></button></th>'
        for key, label in COLUMNS
    )

    body = []
    index = 0
    for cat in categories:
        rows = ""
        for r in cat["records"]:
            index += 1
            badges = ""
            if r["_fresh"]:
                badges += '<span class="badge badge-fresh">Новое</span>'
            if r["relay"]:
                badges += '<span class="badge badge-relay">Эстафета</span>'

            roster = (
                f'<span class="roster">{e(" · ".join(r["roster"]))}</span>'
                if r["roster"] else ""
            )
            rows += (
                f'<tr role="row" data-i="{index}" data-sex="{r["_sex"]}" '
                f'data-pool="{r["_pool"]}" data-stroke="{r["stroke_id"]}" '
                f'data-relay="{"1" if r["relay"] else "0"}" data-date="{r["date"] or ""}" '
                f'data-sec="{r["result_seconds"] or 0}">'
                f'<td role="cell" class="col-disc"><span class="disc-name">{e(r["_title"])}</span>{badges}</td>'
                f'<td role="cell" class="col-athlete">{e(r["athlete"])}{roster}</td>'
                f'<td role="cell" class="col-result" data-label="Результат"><span class="result-value">{e(r["result"])}</span></td>'
                f'<td role="cell" class="col-location" data-label="Место">{e(r["location"])}</td>'
                f'<td role="cell" class="col-date" data-label="Дата">'
                f'<time datetime="{r["date"] or ""}">{fmt_date_ru(r["date"], r["date_original"])}</time></td>'
                "</tr>"
            )

        group = ""
        if len(categories) > 1:
            group = (
                f'<tr class="group-row" role="row"><td role="cell" colspan="5">'
                f'<span class="group-title">{e(cat["title"])}</span></td></tr>'
            )
        body.append(
            f'<tbody role="rowgroup" data-cat="{cat["id"]}" aria-label="{e(cat["title"])}">'
            f'{group}{rows}</tbody>'
        )

    return (
        '<div class="table-frame">'
        '<div class="table-scroll">'
        f'<table class="records" role="table">'
        f'<caption class="sr-only">{e(caption)}</caption>'
        f'<thead role="rowgroup"><tr role="row">{head}</tr></thead>'
        f'{"".join(body)}'
        "</table></div>"
        '<p class="empty-state" hidden>Ничего не найдено. '
        '<button type="button" class="btn-reset">Сбросить</button></p>'
        "</div>"
    )


def render_nav(categories: list[dict], current: str) -> str:
    e = html.escape
    links = [("", "Все рекорды")] + [(c["id"], c["title"]) for c in categories]
    out = ""
    for cid, title in links:
        href = f"/{cid}/" if cid else "/"
        aria = ' aria-current="page"' if cid == current else ""
        cls = "cat-link is-current" if cid == current else "cat-link"
        out += f'<a class="{cls}" href="{href}"{aria}>{e(title)}</a>'
    return f'<nav class="cat-nav" aria-label="Категории рекордов">{out}</nav>'


# ---------------------------------------------------------------- шаблон страницы

# В шаблоне НЕТ ни одной литеральной фигурной скобки: всё, что их содержит
# (инлайновые скрипты, JSON-LD), приходит готовой строкой через плейсхолдер.
THEME_SCRIPT = """
(function () {
  var d = document.documentElement;
  d.classList.remove("no-js");
  var lik = "obsidian";
  try {
    var saved = localStorage.getItem("lik");
    if (saved === "titan" || saved === "obsidian") lik = saved;
    else if (window.matchMedia("(prefers-color-scheme: light)").matches) lik = "titan";
  } catch (e) {}
  d.setAttribute("data-theme", lik);
  var m = document.getElementById("theme-color");
  if (m) m.setAttribute("content", lik === "titan" ? "#fafafa" : "#0d0d0d");
})();
"""

# Стаб ym() и сам init() создаются сразу (синхронно): app.js шлёт цели (reachGoal)
# с первого взаимодействия пользователя, и без стаба ранние вызовы молча терялись бы
# до window.load. Тяжёлый tag.js с вебвизором (~107K gzip) по-прежнему грузится лениво
# после load — он конкурировал со шрифтами и CSS за первый рендер; очередь вызовов,
# накопленная стабом, применяется сама, как только tag.js подгрузится.
METRIKA = """
(function(m,e,t,r,i,k,a){
    m[i]=m[i]||function(){(m[i].a=m[i].a||[]).push(arguments)};
    m[i].l=1*new Date();
})(window, document, 'script', 'https://mc.yandex.ru/metrika/tag.js?id=109048777', 'ym');
window.YM_COUNTER_ID = 109048777;
ym(109048777, 'init', {ssr:true, webvisor:true, clickmap:true, ecommerce:"dataLayer", referrer: document.referrer, url: location.href, accurateTrackBounce:true, trackLinks:true});
(function () {
  function start() {
    for (var j = 0; j < document.scripts.length; j++) {if (document.scripts[j].src === 'https://mc.yandex.ru/metrika/tag.js?id=109048777') { return; }}
    var k = document.createElement('script'), a = document.getElementsByTagName('script')[0];
    k.async = 1;
    k.src = 'https://mc.yandex.ru/metrika/tag.js?id=109048777';
    a.parentNode.insertBefore(k, a);
  }
  if (document.readyState === "complete") start();
  else window.addEventListener("load", start, { once: true });
})();
"""

PAGE_TEMPLATE = """<!doctype html>
<html lang="ru" class="no-js" data-theme="obsidian">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<meta name="theme-color" id="theme-color" content="#0d0d0d">
<script>{theme_script}</script>
<title>{title}</title>
<meta name="description" content="{description}">
<meta name="keywords" content="{keywords}">
<meta name="robots" content="index, follow">
<meta name="author" content="Никита Бороздов">
<link rel="canonical" href="{canonical}">
<meta property="og:type" content="website">
<meta property="og:site_name" content="Юношеские рекорды России по плаванию">
<meta property="og:title" content="{og_title}">
<meta property="og:description" content="{description}">
<meta property="og:url" content="{canonical}">
<meta property="og:locale" content="ru_RU">
<meta property="og:image" content="https://{domain}/og-image.png">
<meta property="og:image:width" content="1200">
<meta property="og:image:height" content="630">
<meta property="og:image:alt" content="Юношеские рекорды России по плаванию — актуальная таблица">
<meta name="twitter:card" content="summary_large_image">
<meta name="twitter:title" content="{og_title}">
<meta name="twitter:description" content="{description}">
<meta name="twitter:image" content="https://{domain}/og-image.png">
<meta name="google-site-verification" content="qOwWmdq24kGcVTyxc1GL2W8TxQk63Z5lBH3NSv4hH4s">
<meta name="yandex-verification" content="80f947e774535d84">
<link rel="icon" type="image/png" sizes="120x120" href="/favicon-120.png">
<link rel="icon" href="/favicon.ico" sizes="48x48">
<link rel="apple-touch-icon" href="/apple-touch-icon.png">
<meta name="apple-mobile-web-app-title" content="Юниоры">
<meta name="mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="default">
<link rel="manifest" href="/site.webmanifest">
{startup_images}<link rel="preload" href="{font_inter_cyrillic}" as="font" type="font/woff2" crossorigin>
<link rel="preload" href="{font_inter_latin}" as="font" type="font/woff2" crossorigin>
<link rel="preload" href="{font_mono_latin}" as="font" type="font/woff2" crossorigin>
<style>{fonts_css}</style>
<link rel="stylesheet" href="{style_url}">
<script>{metrika}</script>
<noscript><div><img src="https://mc.yandex.ru/watch/109048777" style="position:absolute; left:-9999px;" alt=""></div></noscript>
<script type="application/ld+json">{jsonld}</script>
</head>
<body>
<a class="skip-link" href="#table">К таблице рекордов</a>

<header class="site-header">
  <div class="shell header-inner">
    <div class="brand">
      <a class="wordmark" href="/">Рекорды России<span class="wordmark-sub">Плавание · Юниоры</span></a>
    </div>
    <div class="header-actions">
      <button class="icon-btn js-only" id="theme-toggle" aria-label="Сменить лик"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><circle cx="12" cy="12" r="4"/><path d="M12 2v2M12 20v2M4.93 4.93l1.41 1.41M17.66 17.66l1.41 1.41M2 12h2M20 12h2M4.93 19.07l1.41-1.41M17.66 6.34l1.41-1.41"/></svg></button>
      <div class="dl" id="dl-menu">
        <a class="btn btn-primary" id="dl-btn" href="#downloads" aria-haspopup="true" aria-expanded="false" aria-controls="dl-panel">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M12 3v12M6 11l6 6 6-6M4 21h16"/></svg>
          <span class="btn-label">Скачать</span>
        </a>
        <div class="dl-panel" id="dl-panel">
          <span class="label dl-title" aria-hidden="true">Скачать</span>
          <button type="button" class="js-only" id="dl-png-btn"><span class="fmt">PNG</span><span class="hint">картинка таблицы</span></button>
          <button type="button" class="js-only" id="dl-pdf-btn"><span class="fmt">PDF</span><span class="hint">документ таблицы</span></button>
          <span class="dl-sep js-only" role="separator"></span>
{download_menu}          <span class="dl-sep js-only" id="install-sep" role="separator" hidden></span>
          <button type="button" class="js-only" id="install" hidden><span class="fmt">APP</span><span class="hint">установить приложение</span></button>
        </div>
      </div>
    </div>
  </div>
</header>
<p class="net-note label js-only" id="net-note" hidden></p>

<div class="author-strip">
  <div class="shell author-strip-inner">
    <div class="signature">{signature}</div>
    <a class="btn author-link" href="{header_by}" rel="author noopener" target="_blank">borozdov.ru <span class="arrow" aria-hidden="true">&#8594;</span></a>
  </div>
</div>

<main class="shell" id="main">
  <section class="hero">
    {eyebrow_html}
    <h1>{h1}</h1>
    <p class="hero-meta">{intro}</p>
    <div class="stat-strip">{stats_html}</div>
  </section>

  {nav_html}

  <section class="controls js-only" aria-label="Поиск и фильтры">
    <div class="search-row">
      <label class="search">
        <span class="sr-only">Поиск по таблице</span>
        <svg class="search-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" aria-hidden="true"><circle cx="11" cy="11" r="7"/><path d="m21 21-4.3-4.3"/></svg>
        <input id="search" type="search" placeholder="Фамилия, дисциплина, город…" autocomplete="off" autocorrect="off" autocapitalize="off" spellcheck="false" enterkeyhint="search">
      </label>
      <p class="result-count label" id="visible-count" role="status" aria-live="polite">Показано {total} из {total}</p>
    </div>
    <div class="filters" id="filters"></div>
  </section>

  <p class="print-head"><span class="print-title">{h1}</span><span class="print-meta"></span></p>

  <section id="table" tabindex="-1" aria-labelledby="table-h">
    <h2 id="table-h" class="sr-only">{table_h2}</h2>
    {table_html}
  </section>

  {faq_html}
</main>

<footer class="footer">
  <div class="shell footer-inner">
    <div>
      <p>
        Данные синхронизируются раз в сутки с
        <a href="{source_url}" rel="noopener" target="_blank">russwimming.ru</a>.
        Если рекорд отсутствует или выглядит устаревшим — сайт-источник ещё не обновил свою таблицу.
        {history_link}
      </p>
      <p class="footer-downloads" id="downloads">
        Скачать таблицу: {download_links}
      </p>
    </div>
    <div class="footer-exit">
      <div class="signature">{signature}</div>
      <a class="btn author-link" href="{footer_by}" rel="author noopener" target="_blank">borozdov.ru <span class="arrow" aria-hidden="true">&#8594;</span></a>
    </div>
  </div>
  <div class="print-signature"><span class="print-date"></span></div>
</footer>

<button type="button" class="to-top js-only" id="to-top" aria-label="Наверх" hidden>
  <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M12 19V5M5 12l7-7 7 7"/></svg>
</button>
{export_html}
<div class="toast" id="toast"><span class="toast-text" role="status" aria-live="polite"></span></div>
<script src="{app_url}" defer></script>
</body>
</html>
"""


def export_dialog() -> str:
    """Окно настройки картинки. Канва читает содержимое из DOM страницы,
    поэтому сюда передаём только то, чего в DOM нет: адрес для QR."""
    def seg(name, options, checked):
        out = ""
        for value, label in options:
            attr = " checked" if value == checked else ""
            out += (f'<label class="seg-option"><input type="radio" name="{name}" '
                    f'value="{value}"{attr}><span>{label}</span></label>')
        return f'<div class="seg" role="radiogroup">{out}</div>'

    def check(cid, label, checked=True):
        attr = " checked" if checked else ""
        return (f'<label class="export-check"><input type="checkbox" id="{cid}"{attr}>'
                f'<span>{label}</span></label>')

    return f"""
<div class="export js-only" id="export" hidden data-qr="https://{SITE_DOMAIN}/go/">
  <div class="export-scrim" data-close></div>
  <div class="export-panel" role="dialog" aria-modal="true" aria-labelledby="export-h">
    <div class="export-head">
      <h2 id="export-h">Картинка таблицы</h2>
      <button type="button" class="icon-btn" id="export-close" aria-label="Закрыть">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" aria-hidden="true"><path d="M18 6 6 18M6 6l12 12"/></svg>
      </button>
    </div>

    <div class="export-body">
      <div class="export-form">
        <div class="export-field"><span class="label">Лик</span>
          {seg("ex-lik", [("obsidian", "Обсидиан"), ("titan", "Титан")], "obsidian")}</div>
        <div class="export-field"><span class="label">Форма</span>
          {seg("ex-shape", [("page", "Страница"), ("ribbon", "Лента")], "page")}</div>
        <div class="export-field"><label class="label" for="ex-title">Заголовок</label>
          <input type="text" id="ex-title" maxlength="60" autocomplete="off" enterkeyhint="done"></div>
        <div class="export-field"><span class="label">Показывать</span>
          {check("ex-stats", "Статистику")}
          {check("ex-badges", "Бейджи «Новое»")}
          {check("ex-roster", "Состав эстафет")}
        </div>
      </div>

      <div class="export-preview">
        <div class="export-canvas-wrap"><canvas id="export-canvas"></canvas></div>
        <p class="export-note label" id="export-note" role="status" aria-live="polite"></p>
      </div>
    </div>

    <div class="export-actions">
      <button type="button" class="btn" id="export-png">Скачать PNG</button>
      <button type="button" class="btn btn-primary" id="export-pdf">Скачать PDF</button>
    </div>
  </div>
</div>
"""


def stat(value: str, label: str) -> str:
    return (
        f'<div class="stat"><span class="stat-value">{html.escape(value)}</span>'
        f'<span class="label">{html.escape(label)}</span></div>'
    )


def build_stats(records: list[dict], data: dict, fetched_human: str) -> str:
    fresh = sum(1 for r in records if r["_fresh"])
    return (
        stat(str(len(records)), "Действующих рекордов")
        + stat(str(fresh), "Обновлено за 12 месяцев")
        + stat(fetched_human, "Синхронизировано")
    )


def faq_section() -> str:
    # Ответ вставляется как есть: в FAQ_ITEMS лежит HTML со ссылками, и это не
    # пользовательский ввод. Вернуть сюда html.escape — значит показать людям
    # <a href=…> текстом. Вопрос по-прежнему эскейпим.
    items = "".join(
        f'<div class="faq-item"><h3>{html.escape(q)}</h3><p>{a}</p></div>'
        for q, a in FAQ_ITEMS
    )
    return (
        '<section class="content-section" aria-labelledby="faq-h">'
        '<h2 id="faq-h">Вопросы и ответы</h2>'
        f'<div class="faq-grid">{items}</div></section>'
    )


def jsonld_graph(data: dict, canonical: str, name: str, description: str,
                 records: list[dict], with_faq: bool) -> str:
    years = [r["date"][:4] for r in records if r["date"]]
    year_min, year_max = (min(years), max(years)) if years else ("", "")
    graph = [
        {
            "@type": "WebSite",
            "@id": f"https://{SITE_DOMAIN}/#website",
            "url": f"https://{SITE_DOMAIN}/",
            "name": SITE_TITLE,
            "description": SITE_TAGLINE,
            "inLanguage": "ru",
            "publisher": {"@type": "Person", "name": "Никита Бороздов", "url": "https://borozdov.ru/"},
        },
        {
            "@type": "Dataset",
            "@id": canonical + "#dataset",
            "name": name,
            "description": description,
            "url": canonical,
            "inLanguage": "ru",
            "keywords": ["юношеские рекорды России", "рекорды России", "плавание", "бассейн 50 м", "бассейн 25 м", "эстафеты"],
            "dateModified": data["fetched_at"],
            "temporalCoverage": f"{year_min}/{year_max}",
            "isBasedOn": data["source_url"],
            "license": "https://creativecommons.org/licenses/by/4.0/",
            "creator": {"@type": "Organization", "name": "Всероссийская федерация плавания",
                        "url": "https://russwimming.ru"},
            "distribution": [
                {"@type": "DataDownload", "encodingFormat": fmt_type,
                 "contentUrl": f"https://{SITE_DOMAIN}/{fname}"}
                for fname, fmt_type in [
                    ("records.json", "application/json"),
                    ("records.csv", "text/csv"),
                    ("records.xlsx",
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                    ("records.md", "text/markdown"),
                    ("records.txt", "text/plain"),
                ]
            ],
        },
    ]
    if with_faq:
        graph.append({
            "@type": "FAQPage",
            "@id": f"https://{SITE_DOMAIN}/#faq",
            "mainEntity": [
                {"@type": "Question", "name": q,
                 "acceptedAnswer": {"@type": "Answer", "text": faq_plain(a)}}
                for q, a in FAQ_ITEMS
            ],
        })
    return json.dumps({"@context": "https://schema.org", "@graph": graph}, ensure_ascii=False)


def history_link() -> str:
    """Самостоятельное предложение в конце абзаца про синхронизацию: разделитель
    «·» после точки читался бы мусором. Адрес — repo_url(), одинаковый локально и в CI."""
    return (f'Каждое обновление видно в '
            f'<a href="{repo_url()}/commits/main/data/junior.json" '
            f'rel="noopener" target="_blank">истории изменений</a>.')


def render_page(*, data: dict, categories: list[dict], canonical: str, title: str,
                og_title: str, description: str, h1: str, eyebrow: str, intro: str,
                current: str, fetched_human: str, with_faq: bool,
                table_h2: str, caption: str) -> str:
    records = [r for cat in categories for r in cat["records"]]
    download_menu = "".join(
        f'          <a href="/{fname}" download><span class="fmt">{fmt}</span>'
        f'<span class="hint">{hint}</span></a>\n'
        for fname, fmt, hint in DOWNLOAD_FORMATS
    )
    download_links = " · ".join(
        f'<a href="/{fname}" download>{fmt}</a>' for fname, fmt, _ in DOWNLOAD_FORMATS
    )
    return PAGE_TEMPLATE.format(
        theme_script=THEME_SCRIPT,
        metrika=METRIKA,
        startup_images=startup_images(),
        fonts_css=fonts_css(),
        font_inter_cyrillic=asset_url("fonts/inter-cyrillic.woff2"),
        font_inter_latin=asset_url("fonts/inter-latin.woff2"),
        font_mono_latin=asset_url("fonts/jetbrains-mono-latin.woff2"),
        style_url=asset_url("style.css"),
        app_url=asset_url("app.js"),
        title=html.escape(title),
        og_title=html.escape(og_title),
        description=html.escape(description),
        keywords=html.escape(SITE_KEYWORDS),
        canonical=canonical,
        domain=SITE_DOMAIN,
        jsonld=jsonld_graph(data, canonical, og_title, description, records, with_faq),
        source_url=html.escape(data["source_url"]),
        eyebrow_html=f'<p class="label">{html.escape(eyebrow)}</p>' if eyebrow else "",
        h1=html.escape(h1),
        intro=intro,
        stats_html=build_stats(records, data, fetched_human),
        nav_html=render_nav(data["categories"], current),
        total=len(records),
        table_h2=html.escape(table_h2),
        table_html=render_table(categories, caption),
        faq_html=faq_section() if with_faq else "",
        signature=SIGNATURE,
        header_by=brand_link("https://borozdov.ru/", "header"),
        footer_by=brand_link("https://borozdov.ru/", "footer"),
        history_link=history_link(),
        export_html=export_dialog(),
        download_menu=download_menu,
        download_links=download_links,
    )


def write_index(data: dict, out: Path, fetched_human: str) -> None:
    years = [r["date"][:4] for r in all_records(data) if r["date"]]
    year_max = max(years) if years else ""
    intro = (
        "Действующие юношеские рекорды России во всех дисциплинах — личные и эстафеты, "
        "бассейны 50 и 25 метров. Источник: "
        f'<a href="{html.escape(data["source_url"])}" rel="noopener" target="_blank">russwimming.ru</a>.'
    )
    out.write_text(render_page(
        data=data,
        categories=data["categories"],
        canonical=f"https://{SITE_DOMAIN}/",
        title=f"{SITE_TITLE} — таблица {year_max}, бассейны 50 и 25 м",
        og_title=SITE_TITLE,
        description=SITE_TAGLINE,
        h1="Юношеские рекорды России по плаванию",
        eyebrow="",
        intro=intro,
        current="",
        fetched_human=fetched_human,
        with_faq=True,
        table_h2="Таблица юношеских рекордов России по плаванию",
        caption="Действующие юношеские рекорды России по плаванию по категориям",
    ), encoding="utf-8")


def write_category_pages(data: dict, fetched_human: str) -> list[str]:
    urls = []
    for cat in data["categories"]:
        recs = cat["records"]
        fresh = sum(1 for r in recs if r["_fresh"])
        canonical = f"https://{SITE_DOMAIN}/{cat['id']}/"
        desc = (
            f"{cat['title']}: {len(recs)} действующих юношеских рекордов России по плаванию. "
            f"Обновлено {fetched_human}, за последний год обновлено {fresh}. "
            "Скачать: JSON, CSV, XLSX."
        )
        intro = (
            f"Действующие юношеские рекорды России в категории «{html.escape(cat['title'].lower())}». "
            'Полная таблица со всеми категориями — <a href="/">на главной</a>.'
        )
        page = render_page(
            data=data,
            categories=[cat],
            canonical=canonical,
            title=f"{cat['title']} — юношеские рекорды России по плаванию",
            og_title=f"{cat['title']} — юношеские рекорды России",
            description=desc,
            h1=cat["title"],
            eyebrow="Юношеские рекорды России по плаванию",
            intro=intro,
            current=cat["id"],
            fetched_human=fetched_human,
            with_faq=False,
            table_h2=f"Таблица рекордов: {cat['title']}",
            caption=f"Действующие юношеские рекорды России: {cat['title']}",
        )
        target = PUBLIC / cat["id"] / "index.html"
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(page, encoding="utf-8")
        urls.append(f"/{cat['id']}/")
    return urls


# ---------------------------------------------------------------- служебные файлы

def startup_images() -> str:
    """<link rel="apple-touch-startup-image"> для каждого размера из gen_icons.SPLASH_SIZES.

    Работают только вместе с <meta name="apple-mobile-web-app-capable">. Один
    обсидиановый набор без prefers-color-scheme: сплэш — константа вне зеркала.
    """
    return "".join(
        f'<link rel="apple-touch-startup-image" media="{item["media"]}" href="/{item["name"]}">\n'
        for item in gen_icons.splash_set()
    )


# Порядок иконок в манифесте значим для Android: any-иконки система берёт
# как есть (может подложить белую подложку под непрозрачный PNG), maskable —
# кропит под свою маску. Нужны обе: без any Chrome на части прошивок рисует
# дефолтный глобус, без maskable — квадрат в круглой лунке.
ICON_LETTER = "Ю"


SHORT_SEX = {"women": "Жен.", "men": "Муж.", "mixed": "Микст"}
SHORT_POOL = {"lcm": "50 м", "scm": "25 м"}


def short_title(cat: dict) -> str:
    """Подпись ярлыка ≤ 12 символов: «Жен. 50 м», «Микст»."""
    return f"{SHORT_SEX.get(cat['sex'], '')} {SHORT_POOL.get(cat['pool'], '')}".strip()


def webmanifest(categories: list[dict] | None = None) -> str:
    # UTM-словарь тот же, что у brand_link(): в Метрике запуски с домашнего экрана
    # и с ярлыков видны отдельно от обычных заходов. В JSON `&` пишется как есть.
    return json.dumps({
        "id": f"https://{SITE_DOMAIN}/",
        "name": SITE_TITLE,
        "short_name": "Юниоры",
        "description": SITE_TAGLINE,
        "lang": "ru",
        "dir": "ltr",
        "start_url": "/?utm_source=pwa&utm_medium=homescreen",
        "scope": "/",
        "display": "standalone",
        "display_override": ["standalone", "minimal-ui"],
        "orientation": "any",
        # Обсидиан и в фоне, и в theme_color: сплэш PWA рисуется background_color,
        # и на светлом фоне тёмная иконка вспыхивала бы белым кадром перед стартом.
        "background_color": "#0d0d0d",
        "theme_color": "#0d0d0d",
        "categories": ["sports", "reference"],
        # Повторный запуск с ярлыка переключает в уже открытое окно, а не плодит новые
        "launch_handler": {"client_mode": ["navigate-existing", "auto"]},
        "icons": [
            {"src": "/icon-192.png", "sizes": "192x192", "type": "image/png",
             "purpose": "any"},
            {"src": "/icon-512.png", "sizes": "512x512", "type": "image/png",
             "purpose": "any"},
            {"src": "/icon-maskable-512.png", "sizes": "512x512", "type": "image/png",
             "purpose": "maskable"},
        ],
        # Долгое нажатие на иконку: прямой вход в категорию. Android показывает
        # первые четыре, поэтому порядок — по востребованности, микст последним.
        "shortcuts": [
            {
                "name": cat["title"],
                "short_name": short_title(cat),
                "url": f"/{cat['id']}/?utm_source=pwa&utm_medium=shortcut",
                "icons": [{"src": "/icon-192.png", "sizes": "192x192", "type": "image/png"}],
            }
            for cat in (categories or [])
        ],
        # Richer Install UI в Chrome/Edge: без скриншотов вместо диалога — узкий баннер
        "screenshots": [
            {
                "src": f"/{name}",
                "sizes": f"{size[0]}x{size[1]}",
                "type": "image/png",
                "form_factor": form_factor,
                "label": label,
            }
            for name, form_factor, size, label in gen_screens.SCREENSHOTS
        ],
    }, ensure_ascii=False, indent=2) + "\n"


ROBOTS_TXT = f"""User-agent: *
Allow: /

Sitemap: https://{SITE_DOMAIN}/sitemap.xml
"""

REDIRECT_HTML = f"""<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<title>Юношеские рекорды России по плаванию</title>
<meta name="robots" content="noindex, follow">
<meta http-equiv="refresh" content="0; url={QR_TARGET}">
<link rel="canonical" href="{QR_TARGET}">
</head>
<body>
<p>Переходим к таблице рекордов. <a href="{QR_TARGET}">Открыть вручную →</a></p>
<script>window.location.replace("{QR_TARGET}")</script>
</body>
</html>
"""

NOT_FOUND_HTML = f"""<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<title>Страница не найдена — Юношеские рекорды России по плаванию</title>
<meta name="robots" content="noindex">
<meta http-equiv="refresh" content="0; url=https://{SITE_DOMAIN}/">
<link rel="canonical" href="https://{SITE_DOMAIN}/">
</head>
<body>
<p>Страница не найдена. <a href="https://{SITE_DOMAIN}/">Перейти на главную →</a></p>
<script>window.location.replace("https://{SITE_DOMAIN}/")</script>
</body>
</html>
"""


# Офлайн-страница обязана жить одним файлом: style.css и шрифты лежат в кэше, но
# полагаться на это нельзя — фолбэк показывается ровно тогда, когда что-то пошло
# не так. Значения токенов продублированы из style.css, остальное — только роли.
# Обычная строка, не f-string: внутри CSS с фигурными скобками; подстановка через replace.
OFFLINE_HTML = """<!doctype html>
<html lang="ru" class="no-js" data-theme="obsidian">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<meta name="theme-color" id="theme-color" content="#0d0d0d">
<script>__THEME__</script>
<title>Нет сети — Юношеские рекорды России по плаванию</title>
<meta name="robots" content="noindex">
<style>
:root,[data-theme=obsidian]{color-scheme:dark;--canvas:#0d0d0d;--ink:#fafafa;--slate:#8a8a8a;--strong:#6b6b6b;--hairline:#2e2e2e}
[data-theme=titan]{color-scheme:light;--canvas:#fafafa;--ink:#0d0d0d;--slate:#6b6b6b;--strong:#8a8a8a;--hairline:#e4e4e4}
*{box-sizing:border-box;margin:0}
body{min-height:100dvh;display:flex;flex-direction:column;background:var(--canvas);color:var(--ink);font:16px/1.5 Inter,ui-sans-serif,system-ui,-apple-system,"Segoe UI",Roboto,sans-serif;-webkit-font-smoothing:antialiased;padding:max(28px,env(safe-area-inset-top)) max(20px,env(safe-area-inset-right)) max(24px,env(safe-area-inset-bottom)) max(20px,env(safe-area-inset-left))}
.label{font-size:11px;font-weight:500;letter-spacing:.12em;text-transform:uppercase;color:var(--slate)}
h1{font-size:28px;line-height:1.1;letter-spacing:-.02em;text-transform:uppercase;margin:12px 0 12px;max-width:16ch}
p{color:var(--slate);max-width:38ch}
.row{display:flex;flex-wrap:wrap;gap:10px;margin-top:26px}
.btn{display:inline-flex;align-items:center;justify-content:center;min-height:44px;padding:0 18px;font:inherit;font-size:11px;font-weight:500;letter-spacing:.08em;text-transform:uppercase;color:var(--ink);background:transparent;border:1px solid var(--strong);border-radius:4px;text-decoration:none;cursor:pointer;-webkit-tap-highlight-color:transparent}
.btn-primary{background:var(--ink);color:var(--canvas);border-color:var(--ink)}
footer{margin-top:auto;padding-top:24px;border-top:1px solid var(--hairline)}
</style>
</head>
<body>
<main>
<p class="label">Нет сети</p>
<h1>Эта страница ещё не сохранена</h1>
<p>Таблица доступна без сети после первого открытия. Главная и категории, которые вы уже открывали, работают и сейчас.</p>
<div class="row"><button class="btn btn-primary" type="button" onclick="location.reload()">Повторить</button><a class="btn" href="/">На главную</a></div>
</main>
<footer><span class="label">By Borozdov · __DOMAIN__</span></footer>
</body>
</html>
"""


def offline_html() -> str:
    return OFFLINE_HTML.replace("__THEME__", THEME_SCRIPT).replace("__DOMAIN__", SITE_DOMAIN)


# Service worker целиком состоит из фигурных скобок — подстановка только через replace.
# Два кэша: app-<хеш> меняется вместе с данными (страницы + css + js), static-<хеш> —
# шрифты и vendor, годами тот же: иначе каждое суточное обновление тянуло бы
# 107K шрифтов заново.
SW_TEMPLATE = """/* Service worker — сгенерировано scripts/build.py, руками не править.
   Имена кэшей — хеш содержимого прекэшируемых файлов: версия меняется ровно тогда,
   когда меняются байты, поэтому sw.js стабилен от сборки к сборке. */
"use strict";

const APP = "__APP_CACHE__";
const STATIC = "__STATIC_CACHE__";
const APP_FILES = __APP_FILES__;
const FONTS = __FONTS__;
const OFFLINE = "/offline.html";
const NET_TIMEOUT = 3000;

/* Минуя HTTP-кэш: хостинг не шлёт Cache-Control, браузер кэширует эвристически
   по Last-Modified, и addAll мог бы положить в новый кэш вчерашний style.css
   рядом с сегодняшним HTML */
const fresh = (url) => new Request(url, { cache: "reload" });

self.addEventListener("install", (event) => {
  event.waitUntil((async () => {
    const app = await caches.open(APP);
    await app.addAll(APP_FILES.map(fresh));
    const st = await caches.open(STATIC);
    const have = await Promise.all(FONTS.map((url) => st.match(url)));
    await st.addAll(FONTS.filter((_, i) => !have[i]).map(fresh));
  })());
});

self.addEventListener("activate", (event) => {
  event.waitUntil((async () => {
    const keep = new Set([APP, STATIC]);
    for (const key of await caches.keys()) {
      if (!keep.has(key)) await caches.delete(key);
    }
    if (self.registration.navigationPreload) {
      try { await self.registration.navigationPreload.enable(); } catch (_) {}
    }
    await self.clients.claim();
  })());
});

self.addEventListener("message", (event) => {
  if (event.data && event.data.type === "SKIP_WAITING") self.skipWaiting();
});

const pagePath = (url) => url.pathname.replace(/index\\.html$/, "");

const withTimeout = (promise, ms) => new Promise((resolve, reject) => {
  const timer = setTimeout(() => reject(new Error("timeout")), ms);
  promise.then(
    (value) => { clearTimeout(timer); resolve(value); },
    (err) => { clearTimeout(timer); reject(err); },
  );
});

self.addEventListener("fetch", (event) => {
  const req = event.request;
  if (req.method !== "GET") return;
  const url = new URL(req.url);
  if (url.origin !== self.location.origin) return; // Метрика, borozdov.ru — мимо

  /* Выгрузки: сеть без таймаута (xlsx на медленной сети), при обрыве — последняя
     скачанная копия. Ветка стоит ДО navigate: в Chrome клик по <a download>
     приходит как navigate, и офлайн-страница уезжала бы в файл records.xlsx.
     И ДО поиска по кэшам: иначе после первой загрузки — cache-first навсегда. */
  if (url.pathname.startsWith("/records.")) {
    event.respondWith((async () => {
      const st = await caches.open(STATIC);
      try {
        const res = (await event.preloadResponse) || (await fetch(req));
        if (res.ok) st.put(req, res.clone());
        return res;
      } catch (_) {
        return (await st.match(req)) || Response.error();
      }
    })());
    return;
  }

  /* Страницы: сеть с таймаутом, при обрыве — прекэш этой сборки. Сетевой ответ
     в кэш НЕ кладём: в app-<хеш> лежит только одна сборка целиком. */
  if (req.mode === "navigate") {
    event.respondWith((async () => {
      try {
        return await withTimeout(
          (async () => (await event.preloadResponse) || fetch(req))(),
          NET_TIMEOUT,
        );
      } catch (_) {
        const app = await caches.open(APP);
        const path = pagePath(url);
        return (await app.match(path)) || (await app.match(path + "/")) || app.match(OFFLINE);
      }
    })());
    return;
  }

  event.respondWith((async () => {
    const app = await caches.open(APP);
    const st = await caches.open(STATIC);
    /* Только именованные кэши, никакого глобального поиска по всем: между
       install и activate новой версии он мог бы отдать её файлы к старому HTML */
    const hit = (await app.match(req)) || (await st.match(req));
    if (hit) return hit;
    const res = await fetch(req);
    /* jsPDF и qrcode.js — при первом обращении; хеш static-кэша включает их байты,
       так что обновлённая библиотека сама вытеснит старую копию */
    if (res.ok && url.pathname.startsWith("/assets/vendor/")) st.put(req, res.clone());
    return res;
  })());
});
"""


def fs_path(public: Path, url: str) -> Path:
    """URL прекэша → файл в public/: '/' → index.html, '/x/' → x/index.html, '?v=' отбрасывается."""
    clean = url.split("?", 1)[0]
    rel = clean.lstrip("/")
    return public / (rel + "index.html" if clean.endswith("/") else rel)


def content_hash(public: Path, urls: list[str]) -> str:
    h = hashlib.sha256()
    for url in sorted(urls):  # sorted: порядок обхода ФС недетерминирован
        path = fs_path(public, url)
        if not path.is_file():
            # молча неполный прекэш хуже упавшей сборки: addAll упадёт у всех
            raise FileNotFoundError(f"прекэш: {url} → {path} не существует")
        h.update(url.encode())
        h.update(b"\0")
        h.update(path.read_bytes())
        h.update(b"\0")
    return h.hexdigest()[:12]


def app_files(pages: list[str]) -> list[str]:
    return (["/"] + list(pages)
            + ["/offline.html", asset_url("style.css"), asset_url("app.js"), "/site.webmanifest"])


def service_worker(public: Path, pages: list[str]) -> str:
    """sw.js с именами кэшей от содержимого. Вызывать последним: все файлы уже на диске."""
    files = app_files(pages)
    fonts = font_urls()
    return (SW_TEMPLATE
            .replace("__APP_CACHE__", "app-" + content_hash(public, files))
            .replace("__STATIC_CACHE__", "static-" + content_hash(public, fonts + vendor_urls()))
            .replace("__APP_FILES__", json.dumps(files, ensure_ascii=False))
            .replace("__FONTS__", json.dumps(fonts, ensure_ascii=False)))


def write_sitemap(data: dict, paths: list[str], out: Path) -> None:
    lastmod = data["fetched_at"][:10]
    urls = ""
    for path, priority in [("/", "1.0")] + [(p, "0.8") for p in paths]:
        urls += (
            "  <url>\n"
            f"    <loc>https://{SITE_DOMAIN}{path}</loc>\n"
            f"    <lastmod>{lastmod}</lastmod>\n"
            "    <changefreq>daily</changefreq>\n"
            f"    <priority>{priority}</priority>\n"
            "  </url>\n"
        )
    out.write_text(
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
        f"{urls}</urlset>\n",
        encoding="utf-8",
    )


def main() -> int:
    data = load_data()
    today = datetime.now(timezone.utc).date()
    enrich(data, today)

    fetched_dt = datetime.strptime(data["fetched_at"], "%Y-%m-%dT%H:%M:%SZ").replace(
        tzinfo=timezone.utc)
    fetched_human = fetched_dt.strftime("%d.%m.%Y")

    PUBLIC.mkdir(parents=True, exist_ok=True)
    ASSETS.mkdir(parents=True, exist_ok=True)

    copy_static()
    write_index(data, PUBLIC / "index.html", fetched_human)
    cat_paths = write_category_pages(data, fetched_human)
    write_sitemap(data, cat_paths, PUBLIC / "sitemap.xml")
    (PUBLIC / "robots.txt").write_text(ROBOTS_TXT, encoding="utf-8")
    # Переходник для QR-кода на картинках. В sitemap не попадает намеренно.
    (PUBLIC / "go").mkdir(exist_ok=True)
    (PUBLIC / "go" / "index.html").write_text(REDIRECT_HTML, encoding="utf-8")
    (PUBLIC / "404.html").write_text(NOT_FOUND_HTML, encoding="utf-8")
    (PUBLIC / "offline.html").write_text(offline_html(), encoding="utf-8")

    clean = public_json(data)
    (PUBLIC / "records.json").write_text(
        json.dumps(clean, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_csv(clean, PUBLIC / "records.csv")
    write_xlsx(clean, PUBLIC / "records.xlsx")
    write_markdown(clean, PUBLIC / "records.md")
    write_txt(clean, PUBLIC / "records.txt")

    # bold=True: «Ю» шире и «круглее» «Р» — при идентичной фактической
    # толщине штриха читается визуально легче. Bold компенсирует оптику,
    # не саму толщину (см. gen_icons.draw_letter).
    gen_icons.render(PUBLIC, ICON_LETTER, bold=True)
    gen_screens.render(
        data, PUBLIC,
        title=SITE_TITLE,
        eyebrow="Официальное зеркало · Всероссийская федерация плавания",
        fresh_days=FRESH_DAYS,
        today=today.isoformat(),
    )
    (PUBLIC / "site.webmanifest").write_text(webmanifest(data["categories"]), encoding="utf-8")

    gen_og.render(
        PUBLIC / "og-image.png",
        title=SITE_TITLE,
        label="By Borozdov",
        fact=f"{data['total_records']} рекордов · {fetched_human}",
        domain=SITE_DOMAIN,
    )

    # README держит те же числа, что и главная: пишем его из тех же данных,
    # чтобы описание репозитория не расходилось с сайтом.
    gen_readme.write(
        data,
        domain=SITE_DOMAIN,
        fetched_human=fetched_human,
        formats=DOWNLOAD_FORMATS,
        repo_url=repo_url(),
    )

    # Последним: имена кэшей считаются по байтам уже записанных файлов
    (PUBLIC / "sw.js").write_text(service_worker(PUBLIC, cat_paths), encoding="utf-8")

    print(f"built {PUBLIC.relative_to(ROOT)}/ — {data['total_records']} records, "
          f"{len(cat_paths) + 1} pages")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
